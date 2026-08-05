import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _rows_to_table_data(rows: list[dict]) -> tuple[list[str], list[list[str]]]:
    if not rows:
        return [], []
    headers = list(rows[0].keys())
    data = [[str(row.get(h, "")) for h in headers] for row in rows]
    return headers, data


def render_pdf(title: str, subtitle: str, rows: list[dict]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title=title)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Paragraph(subtitle, styles["Normal"]), Spacer(1, 12)]

    headers, data = _rows_to_table_data(rows)
    if not headers:
        elements.append(Paragraph("No data matches the current filters.", styles["Normal"]))
    else:
        pretty_headers = [h.replace("_", " ").title() for h in headers]
        table = Table([pretty_headers] + data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c3c2b7")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f8")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def render_excel(title: str, rows: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Report"

    headers, data = _rows_to_table_data(rows)
    if headers:
        pretty_headers = [h.replace("_", " ").title() for h in headers]
        sheet.append(pretty_headers)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in data:
            sheet.append(row)
        for i, _ in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = 20
    else:
        sheet.append(["No data matches the current filters."])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
