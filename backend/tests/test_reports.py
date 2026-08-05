"""Ch.71 (10 report types), Ch.113 (PDF/Excel export), Ch.114 (historical comparison)."""

import io

from openpyxl import load_workbook


def test_technician_cannot_access_reports(client, auth_headers):
    tech1 = auth_headers("tech01")
    assert client.get("/api/reports", headers=tech1).status_code == 403


def test_all_ten_report_types_listed(client, auth_headers):
    chef = auth_headers("chef01")
    response = client.get("/api/reports", headers=chef)
    assert response.status_code == 200
    keys = [t["key"] for t in response.json()["data"]]
    assert keys == [
        "daily", "weekly", "monthly", "yearly", "technician", "client", "project", "contract",
        "approval", "planning",
    ]


def test_all_eight_intervention_report_types_run(client, auth_headers):
    admin = auth_headers("admin01")
    for report_type in ["daily", "weekly", "monthly", "yearly", "technician", "client", "project", "contract"]:
        response = client.get(f"/api/reports/interventions?type={report_type}", headers=admin)
        assert response.status_code == 200, report_type
        data = response.json()["data"]
        assert data["report_type"] == report_type
        assert data["summary"]["total_interventions"] == len(data["rows"])


def test_unknown_report_type_rejected(client, auth_headers):
    admin = auth_headers("admin01")
    assert client.get("/api/reports/interventions?type=bogus", headers=admin).status_code == 400


def test_client_filter_matches_interventions_endpoint_exactly(client, auth_headers):
    admin = auth_headers("admin01")
    client_id = client.get("/api/clients", headers=admin, params={"page_size": 1}).json()["data"]["items"][0]["id"]

    report = client.get(f"/api/reports/interventions?type=client&client_id={client_id}", headers=admin).json()["data"]
    expected_total = client.get(f"/api/interventions?client_id={client_id}&page_size=1", headers=admin).json()["data"]["total"]
    assert report["summary"]["total_interventions"] == expected_total


def test_daily_and_weekly_default_windows(client, auth_headers):
    from datetime import date

    admin = auth_headers("admin01")
    daily = client.get("/api/reports/interventions?type=daily", headers=admin).json()["data"]
    assert daily["date_from"] == daily["date_to"]

    weekly = client.get("/api/reports/interventions?type=weekly", headers=admin).json()["data"]
    span = (date.fromisoformat(weekly["date_to"]) - date.fromisoformat(weekly["date_from"])).days
    assert span == 6


def test_approval_report(client, auth_headers):
    admin = auth_headers("admin01")
    response = client.get("/api/reports/approval", headers=admin)
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["total_approved"] + data["total_rejected"] == len(data["rows"])


def test_planning_report(client, auth_headers):
    chef = auth_headers("chef01")
    response = client.get("/api/reports/planning", headers=chef)
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["total_planned"] + data["total_cancelled"] == len(data["rows"])


def test_comparison_report(client, auth_headers):
    admin = auth_headers("admin01")
    response = client.get(
        "/api/reports/comparison",
        params={
            "period_a_from": "2026-01-01", "period_a_to": "2026-01-31",
            "period_b_from": "2026-02-01", "period_b_to": "2026-02-28",
        },
        headers=admin,
    )
    assert response.status_code == 200
    data = response.json()["data"]
    assert "period_a" in data and "period_b" in data
    assert data["period_a"]["total_interventions"] >= 0


class TestExports:
    def test_pdf_export_produces_valid_pdf(self, client, auth_headers):
        admin = auth_headers("admin01")
        response = client.get("/api/reports/export.pdf?report=interventions&type=monthly", headers=admin)
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert response.content[:4] == b"%PDF"
        assert len(response.content) > 500

    def test_pdf_export_requires_type_for_interventions(self, client, auth_headers):
        admin = auth_headers("admin01")
        response = client.get("/api/reports/export.pdf?report=interventions", headers=admin)
        assert response.status_code == 400

    def test_pdf_export_of_empty_result_still_valid(self, client, auth_headers):
        admin = auth_headers("admin01")
        response = client.get(
            "/api/reports/export.pdf?report=interventions&type=daily&date_from=1999-01-01&date_to=1999-01-01",
            headers=admin,
        )
        assert response.status_code == 200
        assert response.content[:4] == b"%PDF"

    def test_excel_export_is_genuinely_parseable(self, client, auth_headers):
        admin = auth_headers("admin01")
        response = client.get("/api/reports/export.xlsx?report=interventions&type=technician", headers=admin)
        assert response.status_code == 200
        assert response.content[:2] == b"PK"

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        header_row = [cell.value for cell in sheet[1]]
        assert "Bi Number" in header_row

        expected_rows = len(
            client.get("/api/reports/interventions?type=technician", headers=admin).json()["data"]["rows"]
        )
        assert sheet.max_row - 1 == expected_rows

    def test_technician_cannot_export(self, client, auth_headers):
        tech1 = auth_headers("tech01")
        response = client.get("/api/reports/export.pdf?report=interventions&type=daily", headers=tech1)
        assert response.status_code == 403
